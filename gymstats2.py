#! python3

import openpyxl, time, re, datetime
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


start_time = time.time()

today = datetime.date.today()
yesterday = today - datetime.timedelta(days = 1)
tomorrow = today + datetime.timedelta(days = 1)

today = str(today)
yesterday = str(yesterday)
tomorrow = str(tomorrow)

district_names = ['Austin North', 'Austin South', 'Central Texas']

# 'Colorado', 'Dallas', 'Oklahoma', 'California', 'San Antonio Central', 'San Antonio North', 'San Antonio South']



districts = {'Austin North' :  ['TX-AUSTIN ANDERSON ARBOR',
                      'TX-AUSTIN CEDAR PARK',
                      'TX-AUSTIN CYPRESS CREEK',
                      'TX-AUSTIN HESTERS CROSSING',
                      'TX-AUSTIN NORTH ROUND ROCK',
                      'TX-AUSTIN TECHRIDGE',
                      'TX-GEORGETOWN',
                      'TX-PFLUGERVILLE'],
             'Austin South' : ['TX-AUSTIN BEE CAVES',
                      'TX-AUSTIN BELTERRA',
                      'TX-AUSTIN DOWNTOWN 6THBRAZOS',
                      'TX-AUSTIN HIGHLAND',
                      'TX-AUSTIN NORTH',
                      'TX-AUSTIN SOUTH',
                      'TX-AUSTIN SOUTH CENTRAL',
                      'TX-AUSTIN SOUTHEAST',
                      'TX-AUSTIN WESTLAKE',
                      'TX-SAN MARCOS'],
             'Central Texas' : ['TX-BELLMEAD',
                      'TX-BRYAN',
                      'TX-COLLEGE STATION',
                      'TX-COPPERAS COVE',
                      'TX-KILLEEN',
                      'TX-TEMPLE',
                      'TX-TOWER POINT',
                      'TX-VICTORIA',
                      'TX-WACO'],
             'Colorado' : ['CO-BRIARGATE',
                      'CO-RUSTIC HILLS',
                      'CO-SECURITY'],
             'Dallas' : ['TX-DALLAS UPTOWN',
                      'TX-DESOTO',
                      'TX-PRESTON CENTER',
                      'TX-RICHARDSON',
                      'TX-WAXAHACHIE'],
             'Oklahoma' : ['OK-DEL CITY',
                     'OK-EDMUND',
                     'OK-MOORE',
                     'OK-NORMAN',
                     'OK-NORTHWEST',
                     'OK-PENN CROSSING',
                     'OK-QUAIL SPRINGS',
                     'OK-WEST EDMUND',
                     'OK-YUKON',
                     'TX-WICHITA FALLS'],
             'California' : ['CA-LONG BEACH',
                      'CA-PASADENA',
                      'CA-VENICE'],
             'San Antonio Central' : ['TX-SAN ANTONIO BANDERA POINTE',
                      'TX-SAN ANTONIO BANDERA TRAILS',
                      'TX-SAN ANTONIO MEDICAL CENTER',
                      'TX-SAN ANTONIO TEZEL'],
             'San Antonio North' : ['TX-NEW BRAUNFELS',
                      'TX-SAN ANTONIO 78THWALZEM',
                      'TX-SAN ANTONIO ALAMO HEIGHTS',
                      'TX-SAN ANTONIO BULVERDE',
                      'TX-SAN ANTONIO EVANS ROAD',
                      'TX-SAN ANTONIO HILL COUNTRY VILLAGE',
                      'TX-SAN ANTONIO LIVE OAK',
                      'TX-SAN ANTONIO LOCKHILL VILLAGE',
                      'TX-SAN ANTONIO NACO',
                      'TX-SAN ANTONIO THE QUARRY',
                      'TX-SAN ANTONIO TRAVIS'],
             'San Antonio South' : ['TX-BROWNSVILLE',
                      'TX-SAN ANTONIO BABCOCK',
                      'TX-SAN ANTONIO CROSSROADS',
                      'TX-SAN ANTONIO CULEBRA',
                      'TX-SAN ANTONIO FIESTA TRAILS',
                      'TX-SAN ANTONIO GOLIAD',
                      'TX-SAN ANTONIO LEGACY TRAILS',
                      'TX-SAN ANTONIO MILITARY',
                      'TX-SAN ANTONIO PRUE ROAD',
                      'TX-SAN ANTONIO ROGERS RANCH',
                      'TX-SAN ANTONIO VALLEY HI',
                      'TX-SOUTH STAPLES']
             }


def loadGym(gyms):
    dictionary = {}
    for i in range(len(gyms)):
        dictionary.update({gyms[i] : {}})
    return dictionary

def lastFirst(first_last):
    if first_last == 'Marziyeh Nasiri Shoja':
        last_first = 'Nasiri Shoja, Marziyeh'
    else:
        name = first_last.split(' ')
        last = name.pop()
        name.insert(0, last)
        name[0] = name[0] + ','
        last_first = ' '.join(name)
    return last_first

def noDollar(number):
    if number:
        number = str(number)
        number = float(number.replace('$', ''))
    else:
        number = 0
    return number

def districtStats(district, district_gyms):

    print(district)
    GYMS = loadGym(district_gyms)
    SALES = loadGym(district_gyms)
    PT = loadGym(district_gyms)

    for v in GYMS.values():
        v.update( {'POS NBU' : 0, 'NMC' : 0, 'POS FP Set' : 0, 'POS AA' : 0,
                   'Sessions' : 0, 'Clients': 0, 'Frequency': 0,
                   'APOS NBU' : 0, 'POS Rev' : 0, 'APOS Rev' : 0, 'FP Set' : 0, 'FP Show' : 0,} )

    print('Calculating Gym and Membership Data...')
    GGEWB = openpyxl.load_workbook('Golds Gym Experience Report with Detail.xlsx') 
    ggesheet = GGEWB['New Membership Agreements Detai']

    gym_name = column_index_from_string('A')
    isNM = column_index_from_string('V')
    isDB = column_index_from_string('G')
    isRet = column_index_from_string('R')
    isFAO = column_index_from_string('J')
    FP_y = column_index_from_string('X')
    AA_tier = column_index_from_string('K')
    sales_person = column_index_from_string('AE')
    noMem = ['Staff', 'Trade', 'Lead']

    for row in ggesheet.rows:
        gym = row[gym_name-1].value
        NM = row[isNM-1].value
        DB = row[isDB-1].value
        Ret = row[isRet-1].value
        FAO = row[isFAO-1].value
        tier = row[AA_tier-1].value
        salesperson = row[sales_person-1].value
        FP = row[FP_y-1].value
        if gym in GYMS and NM:
            if DB.startswith('z') or any(x in (DB) for x in noMem):     
                continue
            else:
                GYMS[gym]['NMC'] += 1
                if SALES.get(gym, {}).get(salesperson):
                    SALES[gym][salesperson]['NMC'] += 1
                    if 'Retail' in FAO:
                        SALES[gym][salesperson]['Mem Unit'] += 1
                    else:
                        SALES[gym][salesperson]['Mem Unit'] += 0.75   
                else:
                    if 'Retail' in FAO:
                        SALES[gym].update( {salesperson : {'NMC' : 1, 'FP' : 0, 'AA' : 0, 'NBU' : 0, 'Rev' : 0,
                                                           'Mem Unit' : 0.75, 'AA Unit' : 0}} )
                    else:
                        SALES[gym].update( {salesperson : {'NMC' : 1, 'FP' : 0, 'AA' : 0, 'NBU' : 0, 'Rev' : 0,
                                                           'Mem Unit' : 1.0, 'AA Unit' : 0}} )
            if FP:
                SALES[gym][salesperson]['FP'] += 1
                GYMS[gym]['POS FP Set'] += 1
            if tier != 'Access':
                SALES[gym][salesperson]['AA'] += 1
                GYMS[gym]['POS AA'] += 1
                if tier != 'GoldsPTx/Mo':
                    if 'Enhanced' in tier:
                        SALES[gym][salesperson]['AA Unit'] += 0.25
                    elif 'Bootcamp' or 'Studio' in tier:
                        SALES[gym][salesperson]['AA Unit'] += 0.5
                    else:
                        continue


    print('Calculating Active Clients...')
    ptSessionsWB = openpyxl.load_workbook('PT Business Report - PT Sessions Serviced with Date Range.xlsx')
    gsheet = ptSessionsWB['PT Sessions Serviced Summary']
    tsheet = ptSessionsWB['PT Sessions Serviced Individual']

    gGymName = column_index_from_string('D')
    gClients = column_index_from_string('F')
    gFreq = column_index_from_string('J')
    #gSes = column_index_from_string('H')
                                              
    tGymName = column_index_from_string('E')
    ptName = column_index_from_string('F')
    tClients = column_index_from_string('G')
    tFreq = column_index_from_string('K')
    #tSes = column_index_from_string('I')

    for row in gsheet.rows:
        gym = row[gGymName-1].value
        clients = row[gClients-1].value
        freq = row[gFreq-1].value
        #ses = row[gSes-1].value
        if gym in PT:
            GYMS[gym]['Clients'] = clients
            GYMS[gym]['Frequency'] = freq
            #GYMS[gym]['Sessions'] = ses

    for row in tsheet.rows:
        gym = row[tGymName-1].value
        clients = row[tClients-1].value
        freq = row[tFreq-1].value
        pt = row[ptName-1].value
        #ses = row[tSes-1].value
        if gym in PT:
            if pt in PT[gym]:
                PT[gym][pt]['Clients'] = clients
                #PT[gym][pt]['Sessions'] = sessions
                PT[gym][pt]['Frequency'] = frequency
            else:
                PT[gym].update( {pt : {'Sessions' : 0, 'Classes' : 0, 'Clients' : clients, 'Frequency' : freq,
                                       'Set' : 0, 'Show' : 0, 'NBU' : 0, 'PT New Rev': 0, 'PT Renew Rev': 0} })      #if they're not listed, add them with some stats
                        

    print('Calculating Sessions Serviced...')
    ptTrainingWB = openpyxl.load_workbook('PT Training Payroll Report.xlsx')
    pt_sheet = ptTrainingWB['PT_Payroll_Summary']
                                              
    pt_gym_name = column_index_from_string('G')
    pt_name = column_index_from_string('E')
    pt_ses = column_index_from_string('H')

    for row in pt_sheet.rows:
          gym = row[pt_gym_name-1].value
          ses = row[pt_ses-1].value
          pt = row[pt_name-1].value
          if gym in PT:
              GYMS[gym]['Sessions'] += ses
              if pt in PT[gym]:
                  PT[gym][pt]['Sessions'] = ses
              else:
                  PT[gym].update( {pt : {'Sessions' : ses, 'Classes' : 0, 'Clients' : clients, 'Frequency' : freq,
                                         'Set' : 0, 'Show' : 0, 'NBU' : 0, 'PT New Rev': 0, 'PT Renew Rev': 0} })      #if they're not listed, add them with some stats
              

    print('Calculating FP...')
    ptfpWB = openpyxl.load_workbook('Service Provider Activity Summary.xlsx')
    fpsheet = ptfpWB['Sheet1']

    gymName = column_index_from_string('A')
    service_provider = column_index_from_string('B')
    app_show = column_index_from_string('I')
    app_service = column_index_from_string('H')
    fp_type = ['GOLD\'S 3D', 'Fitness Profile', 'Fitness Profile Follow-Up', 'Fit Profile', 'Fit Profile Follow-Up',
               'Fitness Assessment']

    for row in fpsheet.rows:
        gym = row[gymName-1].value
        provider = row[service_provider-1].value
        show = row[app_show-1].value
        service = row[app_service-1].value
        if gym in PT and provider != 'No Service Provider' and provider and any(x in (service) for x in fp_type):
            pt = lastFirst(provider)
            GYMS[gym]['FP Set'] += 1
            if pt in PT[gym]:
                PT[gym][pt]['Set'] += 1
            else:
                PT[gym].update( {pt : {'Sessions' : 0, 'Classes' : 0, 'Clients' : 0, 'Frequency' : 0,
                                             'Set' : 1, 'Show' : 0, 'NBU' : 0, 'PT New Rev': 0, 'PT Renew Rev': 0} })      
            if pt in PT[gym] and show:
                PT[gym][pt]['Show'] += 1
                GYMS[gym]['FP Show'] += 1
            
    ptSalesWB = openpyxl.load_workbook('PT Business Report - PT Sales.xlsx')
    sheet = ptSalesWB['PT Business Report - PT Sales']

    gym_name = column_index_from_string('D')
    nbuColumn = column_index_from_string('AG')
    sales_person = column_index_from_string('F')
    department_col = column_index_from_string('G')
    package_name = column_index_from_string('L')
    pos_y = column_index_from_string('S')
    billed = column_index_from_string('W')
    amount = column_index_from_string('X')
    last_act = column_index_from_string('AD')
    pt_sale_date = column_index_from_string('P')
    m_sale_date = column_index_from_string('Q')
    pt_department = ['Asst Fitness Manager', 'Fitness Advisor', 'Fitness Director', 'Fitness Svc Manager 1', 'Fitness Svc Manager 2', 'Fitness Svc Manager 3', 'PT Level 1', 'PT Level 2', 'PT Level 3', 'PT Level 4', 'Studio Coach']
    sales_department = ['Membership Advisor', 'Asst General Manger', 'DM/SVP', 'Front Desk Associate', 'Front Desk Manager', 'General Manager']

    for row in sheet.rows:                                                          #look through each row in the sheet's rows  
        gym = row[gym_name-1].value
        salesperson = row[sales_person-1].value
        department = row[department_col-1].value
        intro = row[package_name-1].value
        pos = row[pos_y-1].value
        nbu = row[nbuColumn-1].value
        last_pt = row[last_act-1].value
        pt_date = row[pt_sale_date-1].value
        m_date = row[m_sale_date-1].value
        paid = row[amount-1].value
        if gym in GYMS and nbu == 'Y':
            invoice = noDollar(row[billed-1].value)
            if pt_date == m_date:
                GYMS[gym]['POS NBU'] += 1
                GYMS[gym]['POS Rev'] += invoice
            else:
                GYMS[gym]['APOS NBU'] += 1
                GYMS[gym]['APOS Rev'] += invoice
                
            if salesperson in SALES[gym]:
                SALES[gym][salesperson]['NBU'] += 1
                SALES[gym][salesperson]['Rev'] += invoice
                if 'INTRO' in intro:
                    SALES[gym][salesperson]['AA Unit'] += 1
            elif salesperson in PT[gym]:
                PT[gym][salesperson]['NBU'] += 1
                PT[gym][salesperson]['PT New Rev'] += invoice
            else:
                if any(x in (department) for x in pt_department):
                    PT[gym].update( {pt : {'Sessions' : 0, 'Classes' : 0, 'Clients' : 0, 'Frequency' : 0,
                                             'Set' : 0, 'Show' : 0, 'NBU' : 1, 'PT New Rev': invoice, 'PT Renew Rev': 0} })      
                elif any(x in (department) for x in sales_department):
                    SALES[gym].update( {salesperson : {'NMC' : 0, 'FP' : 0, 'AA' : 0, 'NBU' : 1, 'Rev' : invoice,
                                                           'Mem Unit' : 0, 'AA Unit' : 0}} )
                else:
                    continue
        elif gym in GYMS and department and paid:
            if last_pt and last_pt in PT[gym]:
                PT[gym][last_pt]['PT Renew Rev'] += paid
            else:
                PT[gym].update( {last_pt : {'Sessions' : 0, 'Classes' : 0, 'Clients' : 0, 'Frequency' : 0,
                                             'Set' : 0, 'Show' : 0, 'NBU' : 0, 'PT New Rev': 0, 'PT Renew Rev': paid} })
                
            
    print('Calculating Classes...')
    classesWB = openpyxl.load_workbook('Daily Service Provider Scheduler.xlsx')
    csheet = classesWB['Sheet1']           

    club_name = column_index_from_string('C')
    service_provider = column_index_from_string('A')
    event = column_index_from_string('U')
    attendance = column_index_from_string('V')
    studio = ['BOOTCAMP', 'GOLD\'S FIT', 'GOLD\'S CYCLE', 'GOLD\'S CYCLE BEATS', 'GOLD\'S CYCLE', 'STUDIO FUSION', 'GOLD\'S BURN']

    for row in csheet.rows:
        gym = row[club_name-1].value
        instructor = row[service_provider-1].value
        classes = row[event-1].value
        attendees = row[attendance-1].value
        if gym in GYMS:
            if instructor in PT[gym] and classes:
                if any(x in (classes) for x in studio) and attendees > 0:
                    PT[gym][instructor]['Classes'] += 1
                else:
                    PT[gym][instructor]['Classes'] += 1

    for gym in GYMS.keys():
        if GYMS[gym]['FP Set'] > 0:
            GYMS[gym].update( {'FP Show %' : GYMS[gym]['FP Show']/GYMS[gym]['FP Set'] })
            GYMS[gym].update( {'FP Close %' : GYMS[gym]['APOS NBU']/GYMS[gym]['FP Set'] })
        if GYMS[gym]['NMC'] > 0:
            GYMS[gym].update( {'POS PT Mix' : GYMS[gym]['POS NBU']/GYMS[gym]['NMC'] })
            GYMS[gym].update( {'FP Set %' : GYMS[gym]['POS FP Set']/GYMS[gym]['NMC'] })
            GYMS[gym].update( {'AA %' : GYMS[gym]['POS AA']/GYMS[gym]['NMC'] })
        if GYMS[gym]['POS NBU'] > 0:
            GYMS[gym].update( {'AVG POS' : GYMS[gym]['POS Rev']/GYMS[gym]['POS NBU'] })
        if GYMS[gym]['APOS NBU'] > 0:
            GYMS[gym].update( {'AVG APOS' : GYMS[gym]['APOS Rev']/GYMS[gym]['APOS NBU'] })

    for gym,team in SALES.items():
        for person in team.keys():
            if SALES[gym][person]['NMC'] > 0:
                SALES[gym][person].update( {'FP %' : SALES[gym][person]['FP']/SALES[gym][person]['NMC'] })
                SALES[gym][person].update( {'AA %' : SALES[gym][person]['AA']/SALES[gym][person]['NMC'] })

    for gym,team in PT.items():
        for person in team.keys():
            if PT[gym][person]['Set'] > 0:
                PT[gym][person].update( {'Show %' : PT[gym][person]['Show']/PT[gym][person]['Set'] })
            if PT[gym][person]['Show'] > 0:
                PT[gym][person].update( {'Clase %' : PT[gym][person]['NBU']/PT[gym][person]['Show'] })


    print('Writing...')
    WB = Workbook()
    sheet = WB.active
    WB.remove(sheet)

    WB.create_sheet('District')
    sheet = WB['District']

    #           A-1       B-2    C-3      D-4          E-5      F-6             G-7       H-8          I-9      J-10       K-11         L-12        M-13        N-14        O-15           P-16       Q-17            R-18       S-19               U-21     V-22
    headers = ['Gym', 'POS NBU', 'NMC', 'POS FP Set', 'POS AA', 'Sessions',  'Clients', 'Frequency', 'APOS NBU', 'POS Rev', 'APOS Rev', 'FP Set', 'FP Show', 'FP Show %', 'FP Close %','POS PT MIX', 'FP POS Set %', 'AA %', 'Avg POS NBU', 'Avg APOS NBU']

    for i in range(len(headers)):
        sheet.cell(row=1, column=i+1).font = Font(bold=True)
        sheet.cell(row=1, column=i+1).value = headers[i]

    row = 2
    for gym,stats in GYMS.items():
        sheet.cell(row=row, column=1, value=gym)
        column = 2
        for stats, numbers in stats.items():
            sheet.cell(row=row, column=column, value=numbers)
            column += 1
        row += 1


    WB.create_sheet('Sales')
    sheet = WB['Sales']

        #             A-1       B-2          C-3    D-4  E-5      F-6       G-7       H-8         I-9      J-10     K-11 
    sales_headers = ['Gym', 'Sales Person', 'NMC', 'FP', 'AA', 'PT NBU', 'Revenue', 'Mem Unit', 'AA Unit', 'FP %', 'AA %']

    for i in range(len(sales_headers)):
        sheet.cell(row=1, column=i+1).font = Font(bold=True)
        sheet.cell(row=1, column=i+1).value = sales_headers[i]

    row = 2
    for gym,team in SALES.items():  
        for salesperson,stats in team.items(): 
            sheet.cell(row=row, column=1, value=gym)
            sheet.cell(row=row, column=2, value=salesperson)
            column = 3
            for stat,num in stats.items(): 
                sheet.cell(row=row, column=column, value=num) 
                column += 1
            row += 1
        for i in range(len(sales_headers)):
            sheet.cell(row=row, column=i+1).font = Font(bold=True)
            sheet.cell(row=row, column=i+1).value = sales_headers[i]
        row += 1

    WB.create_sheet('PT')
    sheet = WB['PT']

    #               A-1       B-2        C-3      D-4       E-5           F-6       G-7     H-8     I-9        J-10      K-11      L-12      M-13
    pt_headers = ['Gym', 'PT Name', 'Sessions', 'Classes', 'Clients', 'Frequency', 'Set', 'Show', 'Close', 'New Rev', 'Renew Rev', 'Show %', 'Close %']

    for i in range(len(pt_headers)):
        sheet.cell(row=1, column=i+1).font = Font(bold=True)
        sheet.cell(row=1, column=i+1).value = pt_headers[i]

    row = 2
    for gym,team in PT.items():
        for pt,stats in team.items():
            sheet.cell(row=row, column=1, value=gym)
            sheet.cell(row=row, column=2, value=pt)
            column = 3
            for stat,num in stats.items(): 
                sheet.cell(row=row, column=column, value=num) 
                column += 1
            row += 1
        for i in range(len(pt_headers)):
                sheet.cell(row=row, column=i+1).font = Font(bold=True)
                sheet.cell(row=row, column=i+1).value = pt_headers[i]
        row += 1


    print('Saving...')
    file_name = 'stats_' + str(district) + '.xlsx'
    WB.save(file_name)
    WB.close()
    print('Done.')

file_num = 1
for i in range(len(district_names)):
    districtStats(district_names[i], districts[district_names[i]])
    print("--- %s seconds ---" % (time.time() - start_time))
    


